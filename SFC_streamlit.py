from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from io import BytesIO
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


st.set_page_config(page_title="製令排程表核對小工具", page_icon="🔍", layout="wide")

FIELD_ALIASES = {
    "製令": ("製令", "製令號碼", "製令編號"),
    "客戶": ("客戶", "客戶名稱"),
    "P/N": ("P/N", "PN", "料號", "品號"),
    "發料日": ("發料日", "發料日期"),
    "入庫日": ("入庫日", "入庫日期"),
}


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def normalized(value: object) -> str:
    return clean_text(value).upper()


def find_header_row(raw: pd.DataFrame) -> int:
    best_row, best_score = 0, -1
    for row_index in range(min(30, len(raw))):
        values = {normalized(value) for value in raw.iloc[row_index].tolist()}
        score = sum(
            any(normalized(alias) in values for alias in aliases)
            for aliases in FIELD_ALIASES.values()
        )
        if score > best_score:
            best_row, best_score = row_index, score
    if best_score < 2:
        raise ValueError("找不到標題列，請確認排程表內有『製令』及『發料日』欄位。")
    return best_row


def find_column(columns: list[object], field: str) -> object | None:
    aliases = {normalized(alias) for alias in FIELD_ALIASES[field]}
    for column in columns:
        if normalized(column) in aliases:
            return column
    return None


def infer_reference_date(df: pd.DataFrame, columns: list[object]) -> date:
    years: list[int] = []
    dates: list[pd.Timestamp] = []
    for field in ("發料日", "入庫日"):
        column = find_column(columns, field)
        if column is None:
            continue
        parsed = pd.to_datetime(df[column], errors="coerce").dropna()
        dates.extend(parsed.tolist())
        years.extend(parsed.dt.year.astype(int).tolist())
    if dates:
        common_year = Counter(years).most_common(1)[0][0]
        middle = sorted(dates)[len(dates) // 2]
        return date(common_year, middle.month, middle.day)
    return date.today()


def parse_header_date(value: object, reference: date) -> date | None:
    if pd.isna(value) or clean_text(value) == "":
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = re.sub(r"\.\d+$", "", clean_text(value))
    match = re.fullmatch(r"(?:(\d{4})[/-])?(\d{1,2})[/-](\d{1,2})", text)
    if not match:
        return None

    year_text, month_text, day_text = match.groups()
    month, day = int(month_text), int(day_text)
    if year_text:
        try:
            return date(int(year_text), month, day)
        except ValueError:
            return None

    candidates: list[date] = []
    for year in (reference.year - 1, reference.year, reference.year + 1):
        try:
            candidates.append(date(year, month, day))
        except ValueError:
            pass
    return min(candidates, key=lambda item: abs((item - reference).days)) if candidates else None


def analyze_schedule(excel_source, sheet_name: str | int = 0) -> pd.DataFrame:
    raw = pd.read_excel(excel_source, sheet_name=sheet_name, header=None)
    header_row = find_header_row(raw)
    excel_source.seek(0)
    df = pd.read_excel(excel_source, sheet_name=sheet_name, header=header_row)
    columns = list(df.columns)

    order_column = find_column(columns, "製令")
    issue_column = find_column(columns, "發料日")
    if order_column is None or issue_column is None:
        raise ValueError("找不到『製令』或『發料日』欄位。")

    reference_date = infer_reference_date(df, columns)
    date_columns = [
        (column, parsed)
        for column in columns
        if (parsed := parse_header_date(column, reference_date)) is not None
    ]
    if not date_columns:
        raise ValueError("找不到日期欄位，例如 8/24、8/25 或 2026/8/24。")

    output_rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        order_no = clean_text(row.get(order_column))
        if not order_no:
            continue

        in_dates = sorted({
            column_date
            for column, column_date in date_columns
            if normalized(row.get(column)) == "IN"
        })

        if not in_dates:
            status = "⚪ 尚未安排"
        elif len(in_dates) == 1:
            status = "🟢 已安排"
        else:
            status = "🟠 已重排"

        issue_source = find_column(columns, "發料日")
        warehouse_source = find_column(columns, "入庫日")
        output_rows.append({
            "製令": order_no,
            "發料日": row.get(issue_source) if issue_source is not None else "",
            "原IN日": in_dates[0] if in_dates else None,
            "料件IN日": in_dates[-1] if in_dates else None,
            "入庫日": row.get(warehouse_source) if warehouse_source is not None else "",
            "IN安排狀態": status,
        })

    result = pd.DataFrame(output_rows)
    if result.empty:
        raise ValueError("沒有找到有效的製令資料。")

    for column in ("發料日", "原IN日", "料件IN日", "入庫日"):
        result[column] = pd.to_datetime(result[column], errors="coerce")

    status_order = pd.Categorical(
        result["IN安排狀態"],
        categories=["⚪ 尚未安排", "🟠 已重排", "🟢 已安排"],
        ordered=True,
    )
    return (
        result.assign(_狀態順序=status_order)
        .sort_values(["_狀態順序", "發料日", "製令"], na_position="last")
        .drop(columns="_狀態順序")
        .reset_index(drop=True)
    )


def make_excel(result: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    result.to_excel(buffer, index=False, sheet_name="IN日期整理結果", engine="openpyxl")
    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook["IN日期整理結果"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    status_fills = {
        "⚪ 尚未安排": PatternFill("solid", fgColor="E7E6E6"),
        "🟢 已安排": PatternFill("solid", fgColor="E2F0D9"),
        "🟠 已重排": PatternFill("solid", fgColor="FCE4D6"),
    }
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    header_map = {cell.value: cell.column for cell in sheet[1]}
    for column_name in ("發料日", "原IN日", "料件IN日", "入庫日"):
        column_index = header_map.get(column_name)
        if column_index:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row_index, column_index).number_format = "yyyy/mm/dd"

    status_column = header_map.get("IN安排狀態")
    if status_column:
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row_index, status_column)
            if cell.value in status_fills:
                cell.fill = status_fills[cell.value]

    widths = [18, 13, 13, 13, 13, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


st.title("🔍製令排程表核對小工具")
st.caption("上傳「前一日」與「今日」生管排程，系統會自動整理並核對發料日、料件 IN 日與入庫日的異動。")


def normalize_date_value(value) -> pd.Timestamp | None:
    if pd.isna(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    return None if pd.isna(parsed) else parsed.normalize()


def date_text(value) -> str:
    parsed = normalize_date_value(value)
    return "" if parsed is None else parsed.strftime("%Y/%m/%d")


def compare_schedules(previous: pd.DataFrame, current: pd.DataFrame) -> pd.DataFrame:
    """以製令比對前一日與今日排程，顯示真正跨日的異動。"""
    previous_map = previous.drop_duplicates("製令", keep="last").set_index("製令")
    current_map = current.drop_duplicates("製令", keep="last").set_index("製令")

    rows: list[dict[str, object]] = []
    all_orders = list(dict.fromkeys(list(current_map.index) + list(previous_map.index)))

    for order_no in all_orders:
        in_previous = order_no in previous_map.index
        in_current = order_no in current_map.index

        if in_current:
            current_row = current_map.loc[order_no]
            record = current_row.to_dict()
            record["製令"] = order_no
        else:
            previous_row = previous_map.loc[order_no]
            record = previous_row.to_dict()
            record["製令"] = order_no

        changes: list[str] = []

        if not in_previous and in_current:
            change_status = "🆕 今日新增"
        elif in_previous and not in_current:
            change_status = "🗑️ 今日排程已移除"
        else:
            previous_row = previous_map.loc[order_no]
            current_row = current_map.loc[order_no]

            compare_fields = [
                ("發料日", "發料日"),
                ("料件IN日", "IN日"),
                ("入庫日", "入庫日"),
            ]
            for field, label in compare_fields:
                old_value = date_text(previous_row.get(field))
                new_value = date_text(current_row.get(field))
                if old_value != new_value:
                    old_display = old_value or "空白"
                    new_display = new_value or "空白"
                    changes.append(f"{label} {old_display} → {new_display}")

            change_status = "🔄 " + "；".join(changes) if changes else "➖ 無異動"

        record["排程異動"] = change_status
        rows.append(record)

    result = pd.DataFrame(rows)
    preferred = ["製令", "發料日", "原IN日", "料件IN日", "入庫日", "IN安排狀態", "排程異動"]
    for col in preferred:
        if col not in result.columns:
            result[col] = ""
    return result[preferred]


def make_compare_excel(result: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    result.to_excel(buffer, index=False, sheet_name="每日排程核對結果", engine="openpyxl")
    buffer.seek(0)
    workbook = load_workbook(buffer)
    sheet = workbook["每日排程核對結果"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    header_map = {cell.value: cell.column for cell in sheet[1]}
    for column_name in ("發料日", "原IN日", "料件IN日", "入庫日"):
        column_index = header_map.get(column_name)
        if column_index:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row_index, column_index).number_format = "yyyy/mm/dd"

    change_column = header_map.get("排程異動")
    if change_column:
        fills = {
            "🆕": PatternFill("solid", fgColor="E2F0D9"),
            "🗑️": PatternFill("solid", fgColor="E7E6E6"),
            "🔄": PatternFill("solid", fgColor="FFF2CC"),
        }
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row_index, change_column)
            text_value = str(cell.value or "")
            for prefix, fill in fills.items():
                if text_value.startswith(prefix):
                    cell.fill = fill
                    break

    widths = [18, 13, 13, 13, 13, 16, 55]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


col_prev, col_today = st.columns(2)
with col_prev:
    previous_file = st.file_uploader(
        "📂 前一日生管排程",
        type=["xlsx", "xlsm", "xls"],
        key="previous_schedule",
    )
with col_today:
    current_file = st.file_uploader(
        "📂 今日生管排程",
        type=["xlsx", "xlsm", "xls"],
        key="current_schedule",
    )

if previous_file is None or current_file is None:
    st.info("請同時上傳「前一日」與「今日」生管排程，才能核對異動。")
else:
    try:
        previous_bytes = previous_file.getvalue()
        current_bytes = current_file.getvalue()

        previous_excel = pd.ExcelFile(BytesIO(previous_bytes))
        current_excel = pd.ExcelFile(BytesIO(current_bytes))

        col1, col2 = st.columns(2)
        with col1:
            previous_sheet = st.selectbox(
                "前一日工作表",
                previous_excel.sheet_names,
                index=0,
                key="previous_sheet",
            )
        with col2:
            current_sheet = st.selectbox(
                "今日工作表",
                current_excel.sheet_names,
                index=0,
                key="current_sheet",
            )

        if st.button("🔍 開始核對", type="secondary", use_container_width=False):
            with st.spinner("正在整理並核對兩日生管排程..."):
                previous_result = analyze_schedule(BytesIO(previous_bytes), previous_sheet)
                current_result = analyze_schedule(BytesIO(current_bytes), current_sheet)
                compare_result = compare_schedules(previous_result, current_result)

            st.session_state["sfc_compare_result"] = compare_result
            st.session_state["sfc_current_result"] = current_result

        compare_result = st.session_state.get("sfc_compare_result")
        current_result = st.session_state.get("sfc_current_result")

        if compare_result is not None and current_result is not None:
            changed_mask = compare_result["排程異動"].str.startswith("🔄", na=False)
            added_mask = compare_result["排程異動"].str.startswith("🆕", na=False)
            removed_mask = compare_result["排程異動"].str.startswith("🗑️", na=False)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("今日全部製令", len(current_result))
            c2.metric("🔄 今日異動", int(changed_mask.sum()))
            c3.metric("🆕 今日新增", int(added_mask.sum()))
            c4.metric("🗑️ 今日移除", int(removed_mask.sum()))

            important_mask = changed_mask | added_mask | removed_mask
            if important_mask.any():
                st.warning(f"⚠️ 今日共有 {int(important_mask.sum())} 筆排程與前一日不同，請優先確認。")
                st.subheader("⚠️ 今日異動")
                changed_df = compare_result.loc[important_mask].copy()
                for column in ("發料日", "原IN日", "料件IN日", "入庫日"):
                    changed_df[column] = pd.to_datetime(changed_df[column], errors="coerce").dt.strftime("%Y/%m/%d").fillna("")
                st.dataframe(changed_df, use_container_width=True, hide_index=True)
            else:
                st.success("✅ 今日排程與前一日相比，發料日、料件 IN 日及入庫日皆無異動。")

            with st.expander("📋 查看全部核對結果"):
                display_df = compare_result.copy()
                for column in ("發料日", "原IN日", "料件IN日", "入庫日"):
                    display_df[column] = pd.to_datetime(display_df[column], errors="coerce").dt.strftime("%Y/%m/%d").fillna("")
                st.dataframe(display_df, use_container_width=True, hide_index=True)

            excel_bytes = make_compare_excel(compare_result)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                "📥 下載每日排程核對結果",
                data=excel_bytes,
                file_name=f"製令排程核對結果_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
            )
    except Exception as error:
        st.error(f"處理失敗：{error}")
